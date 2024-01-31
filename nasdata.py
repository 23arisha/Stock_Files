from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import datetime
import chromedriver_autoinstaller
import os

# Borders
border_style = 'thin'
border = Border(left=Side(style=border_style), right=Side(style=border_style), top=Side(style=border_style),
                bottom=Side(style=border_style))

# Set up the Selenium service and driver
chromedriver_autoinstaller.install()
options = webdriver.ChromeOptions()
options.add_argument("--incognito")
driver = webdriver.Chrome( options=options)

url = "https:/stockhouse.com/markets/stocks/nasdaq"
driver.get(url)

time.sleep(35)

# CLICK MORE BUTTON
wait = WebDriverWait(driver, 30)
more_elements = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "mhcs-st-more")))

for more_element in more_elements:
    driver.execute_script("arguments[0].click();", more_element)
    time.sleep(2)

# FIND TABLE
tables = driver.find_elements(By.CSS_SELECTOR, "table.mhcs-st")
table = tables[3]

# HEADERS
tr_tag = table.find_element(By.CSS_SELECTOR, "tr.mhcs-st-row.mhcs-header")
header = [i.text for i in tr_tag.find_elements(By.CSS_SELECTOR, 'span.mhcs-header-title')]

index = 1
if index < len(header):
    del header[index]

header.append('Date')

# Create a workbook and a worksheet
workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "52weekhigh"

# Populate the worksheet with headers
sheet1.append(header)
for cell in sheet1[1]:
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)

# DATE
ele = driver.find_element(By.CSS_SELECTOR, "div.mhcs-headers > div:nth-child(2)")
date_str = datetime.datetime.strptime(ele.text.strip().split(",")[0].strip(), '%b %d').strftime('%d-%b')

# Helper function to process table data
def process_table(driver, table, worksheet, date):
    data_tr = table.find_elements(By.CSS_SELECTOR, "tr.mhcs-st-row.mhcs-pointer")
    for j in data_tr:
        td_tag = j.find_elements(By.CSS_SELECTOR, "td.mhcs-st-col")
        data = [x.text for x in td_tag]
        index = 2
        if index < len(data):
            del data[index]
        data.append(date)
        worksheet.append(data[1:])
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")


# Helper function to create worksheet
def create_worksheet(workbook, title):
    ws = workbook.create_sheet(title=title)
    ws.append(header)
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
    return ws

# NASDAQ PERCENT GAINER
print("NDAQ PERCENT GAINER")
process_table(driver, tables[3], sheet1, date_str)

# NASDAQ PERCENT DECLINER
ws_decline = create_worksheet(workbook, "52weeklow")
print("NDAQ PERCENT DECLINER")
process_table(driver, tables[4], ws_decline, date_str)

# NASDAQ Net DECINER
ws_decline = create_worksheet(workbook, "NetDecliner")
print("NDAQ Net DECINER")
process_table(driver, tables[2], ws_decline, date_str)

# NASDAQ Volume Actives
ws_decline = create_worksheet(workbook, "Volume_Actives")
print("NDAQ Volume Actives")
process_table(driver, tables[0], ws_decline, date_str)

# SAVE THE WORKBOOK
output_file = os.path.join(os.path.expanduser("~"), "Downloads", "52week" + date_str.replace('-', '') + ".xlsx")

workbook.save(output_file)



