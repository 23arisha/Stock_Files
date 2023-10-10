from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font ,Color

path = "C:/Users/Selltricks/Downloads/chromedriver-win64/chromedriver-win64/chromedriver.exe"
s = Service(path)
options = webdriver.ChromeOptions()
options.add_argument("--incognito")
options.add_argument('log-level=3')

driver = webdriver.Chrome(service=s, options=options)

symbols=['CYTO' , 'BMR ', 'ASXC' ,' IFRX' ,' GFAI', 'TIO' ,' GRRR' ,'BFLY' ,'LCID' ,'APLI.TO', 'OCTO','MTLO.V','PGY', 'BBAI' ,'HOLO' ,'AAOI' ,'IDAI' ,'RIVN' ,'EVLO' ,'TSLA' , 'MF','GROM' ,'AIXI']

data_previous = []  
data_open = [] 
observe=[]

for symbol in symbols:
    # Construct the URL for each symbol's page
    url = f"https://finance.yahoo.com/quote/{symbol}?p={symbol}"
    driver.get(url)

    # Wait for the page to load completely
    time.sleep(2)  # Adjust the sleep duration as needed

    table = driver.find_element(By.TAG_NAME, "table")
    tr_tag=table.find_elements(By.TAG_NAME, "tr")
    for t in tr_tag:
        td_tag=t.find_elements(By.TAG_NAME, "td")
        data=[x.text for x in td_tag]
         # Filter data and print only "Previous Close" and "Open"
        if data:
            if data[0] == "Previous Close":
                data_previous.append(data[1])

            elif data[0] == "Open":
                data_open.append(data[1])

for p, o in zip(data_previous, data_open):
    if p == o:
        observe.append('Same')
    elif p < o:
        observe.append('Up')
    elif p > o:
        observe.append('Down')

# Close the WebDriver when done
driver.quit()

# EXCEL WORKBOOK



workbook = openpyxl.Workbook()
sheet = workbook.active

down_font= Font(color=Color(rgb="00FF0000"))  # Red color
same_font = Font(color=Color(rgb="000000FF")) #blue
up_font= Font(color=Color(rgb="0000FF00"))  # Green color

# Style settings
border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15

# Insert the headers with styling
date = input("Enter date:")
sheet['A1']=date
sheet['A2'] = 'Symbol'
sheet['B2'] = 'Previous'
sheet['C2'] = 'Open'
sheet['D2'] = 'Observation'
for cell in sheet['2:2']:
    cell.font = bold_font
    cell.alignment = center_alignment
    cell.border = border

for i, symbol in enumerate(symbols):
    row = i + 3
    sheet.cell(row=row, column=1, value=symbol).border = border
    sheet.cell(row=row, column=2, value=data_previous[i]).border = border
    sheet.cell(row=row, column=3, value=data_open[i]).border = border
    sheet.cell(row=row, column=4, value=observe[i]).border = border
    sheet.cell(row=row, column=1, value=symbol).alignment = center_alignment
    sheet.cell(row=row, column=2, value=data_previous[i]).alignment = center_alignment
    sheet.cell(row=row, column=3, value=data_open[i]).alignment = center_alignment
    sheet.cell(row=row, column=4, value=observe[i]).alignment = center_alignment
      # Apply font color based on the observation
    if observe[i] == 'Up':
        sheet.cell(row=row, column=4).font = up_font
    elif observe[i] == 'Down':
        sheet.cell(row=row, column=4).font = down_font
    elif observe[i] == 'Same':
        sheet.cell(row=row, column=4).font = same_font

# Insert two empty rows after 'PGY'
pgy_index = symbols.index('PGY') + 4
sheet.insert_rows(pgy_index, amount=3)

sheet.cell(row=pgy_index + 1, column=1, value='Monitoring').font = bold_font
sheet.cell(row=pgy_index + 1, column=1).alignment = center_alignment
sheet.cell(row=pgy_index + 1, column=1).border = border

# Insert symbols and data after 'Monitoring' with styling
monitoring_symbols = symbols[symbols.index('BBAI'):symbols.index('AIXI') + 1]
monitoring_data_prev = data_previous[symbols.index('BBAI'):symbols.index('AIXI') + 1]
monitoring_data_open = data_open[symbols.index('BBAI'):symbols.index('AIXI') + 1]
monitoring_data_observe = observe[symbols.index('BBAI'):symbols.index('AIXI') + 1]

for i, symbol in enumerate(monitoring_symbols):
    row = pgy_index + 3 + i
    sheet.cell(row=row, column=1, value=symbol).border = border
    sheet.cell(row=row, column=2, value=monitoring_data_prev[i]).border = border
    sheet.cell(row=row, column=3, value=monitoring_data_open[i]).border = border
    sheet.cell(row=row, column=4, value=monitoring_data_observe[i]).border = border
    sheet.cell(row=row, column=4).alignment = center_alignment 
    # Apply font color based on the observation (for the monitoring section)
    if monitoring_data_observe[i] == 'Up':
        sheet.cell(row=row, column=4).font = up_font
    elif monitoring_data_observe[i] == 'Down':
        sheet.cell(row=row, column=4).font = down_font
    elif monitoring_data_observe[i] == 'Same':
        sheet.cell(row=row, column=4).font = same_font

# Save the workbook
output_file = r"C:/Users/Selltricks/Downloads/"+ "Previous_open_" + date + ".xlsx"
workbook.save(output_file)

print("Previous Close:")
print(data_previous)
print("Open:")
print(data_open)